# Primero vamos a importar la libreria que nos permitirá trabajar con archivos de excel.
from openpyxl import Workbook
from openpyxl import load_workbook

#Hacemos una instancia de Workbook
# wb = Workbook()

#Guardamos un archivo con el nombre que deseamos
# wb.save("prueba.xlsx")

wb2 = load_workbook("prueba.xlsx")

#ws1 = wb2.create_sheet("hoja1")
#ws2 = wb2.create_sheet("hoja2")
#ws2 = wb2.create_sheet("hoja3", 0)



wb2.save("prueba.xlsx")