from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def crearDocumento(): 
    #Crear libro de excel y la hoja-------------------------------------------------------------
    libro = Workbook()
    #libro = load_workbook('prueba.xlsx')
    hoja = libro.active

    #Eliminar gráfica si existe anteriormente
    hoja._charts.clear()

    #Asignación de datos para hacer las pruebas--------------------------------------------------
    productos = ['Camisa', 'Blusa', 'Zapatos', 'Lentes', 'Pantalón', 'Otro', 'Gorra']
    precios = [145.75, 99.90, 742.8, 500, 125, 299, 153]
    codigo_de_producto = [1, 2, 3, 4, 5, 6, 7]
    estatus = ['Activo', 'Activo', 'Inactivo', 'Activo', 'Inactivo', 'Activo', 'Inactivo']
    #Valor máximo pues puede haber variables mas grandes que otras
    valor_maximo = max(len(productos), len(precios), len(codigo_de_producto), len(estatus))

    #Diseños a aplicar---------------------------------------------------------------------------
    font_bold = Font(bold=True)
    tituloForma = Font(name='Arial', size=24, bold=True, color="000000")
    estiloTitulo = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
    llenarEncabezadoTabla = PatternFill(start_color="6AA121", end_color="6AA121", fill_type="solid")
    llenarContenidoTabla = PatternFill(start_color="FBB917", end_color="FBB917", fill_type="solid")
    llenarEncabezadoTabla.font = font_bold

    #Se crean los encabezados que tendrá la tabla-------------------------------------------------
    titulo = hoja.cell(row=2, column=9, value="Muestra de productos")
    titulo.fill = estiloTitulo
    titulo.font = tituloForma

    estilo_borde = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
    hoja.cell(row=4, column=3, value="Nombre del producto").fill = llenarEncabezadoTabla
    hoja.cell(row=4, column=4, value="Precio").fill = llenarEncabezadoTabla
    hoja.cell(row=4, column=5, value="Código producto").fill = llenarEncabezadoTabla
    hoja.cell(row=4, column=6, value="Estatus").fill = llenarEncabezadoTabla

    hoja.cell(row=4, column=3, value="Nombre del producto").alignment = Alignment(horizontal='center')
    hoja.cell(row=4, column=4, value="Precio").alignment = Alignment(horizontal='center')
    hoja.cell(row=4, column=5, value="Código producto").alignment = Alignment(horizontal='center')
    hoja.cell(row=4, column=6, value="Estatus").alignment = Alignment(horizontal='center')
    
    # Definir estilo de borde de las celdas-------------------------------------------------------
    hoja.cell(row=4, column=3).border = estilo_borde
    hoja.cell(row=4, column=4).border = estilo_borde
    hoja.cell(row=4, column=5).border = estilo_borde
    hoja.cell(row=4, column=6).border = estilo_borde

    hoja.cell(row=4, column=3).font = font_bold
    hoja.cell(row=4, column=4).font = font_bold
    hoja.cell(row=4, column=5).font = font_bold
    hoja.cell(row=4, column=6).font = font_bold

    #Se comienzan a recorrer los datos y a asignar en la tabla, además se les aplica un poco más de diseño ----------------
    contador = 1
    suma = 0.0
    posicionGrafica = 0

    for i in range(valor_maximo):
        
        if len(productos) < valor_maximo:
            productos.extend([''] * (valor_maximo - len(productos)))

        if len(precios) < valor_maximo:
            precios.extend([0] * (valor_maximo - len(precios)))

        if len(codigo_de_producto) < valor_maximo:
            codigo_de_producto.extend([0] * (valor_maximo - len(codigo_de_producto)))
        
        if len(estatus) < valor_maximo:
            estatus.extend([''] * (valor_maximo - len(estatus)))
        
        precio = precios[i]
        producto = productos[i]
        codigo = codigo_de_producto[i]
        est = estatus[i]

        suma += precios[i]

        #Comenzamos a llenar las celdas
        #Se le suma 5 porque a partir de esa fila quiero que inicie la tabla
        hoja.cell(row = i + 5, column=3, value=producto)
        hoja.cell(row = i + 5, column=4, value=precio)
        hoja.cell(row = i + 5, column=5, value=codigo)
        hoja.cell(row = i + 5, column=6, value=est)

        # Agregar color de fondo a las celdas que tienen información
        hoja.cell(row = i + 5, column=3).fill = llenarContenidoTabla
        hoja.cell(row = i + 5, column=4).fill = llenarContenidoTabla
        hoja.cell(row = i + 5, column=5).fill = llenarContenidoTabla
        hoja.cell(row = i + 5, column=6).fill = llenarContenidoTabla

        # Centrar los datos en las celdas
        hoja.cell(row = i + 5, column=3).alignment = Alignment(horizontal='center')
        hoja.cell(row = i + 5, column=4).alignment = Alignment(horizontal='center')
        hoja.cell(row = i + 5, column=5).alignment = Alignment(horizontal='center')
        hoja.cell(row = i + 5, column=6).alignment = Alignment(horizontal='center')

        # Aplicar estilo de borde a las celdas
        hoja.cell(row = i + 5, column=3).border = estilo_borde
        hoja.cell(row = i + 5, column=4).border = estilo_borde
        hoja.cell(row = i + 5, column=5).border = estilo_borde
        hoja.cell(row = i + 5, column=6).border = estilo_borde

        contador += 1
        if contador > valor_maximo:
            hoja.cell(row = i + 6, column=3, value="Total: ").fill = llenarContenidoTabla
            hoja.cell(row = i + 6, column=3).border = estilo_borde
            hoja.cell(row = i + 6, column=3).alignment = Alignment(horizontal='center')
            hoja.cell(row = i + 6, column=3).font = font_bold

            hoja.cell(row = i + 6, column=4, value = f"${ suma }").fill = llenarContenidoTabla
            hoja.cell(row = i + 6, column=4).border = estilo_borde
            hoja.cell(row = i + 6, column=4).alignment = Alignment(horizontal='center')
            hoja.cell(row = i + 6, column=4).font = font_bold
            posicionGrafica = i + 9

    #Ajustar el tamaño de las celdas ----------------------------------------------------------------
    for columna in hoja.columns:
        longitud_maxima = 0
        # Obtiene la letra de la columna (A, B, C, ...)
        column = columna[0].column_letter  
        for celda in columna:
            try:
                if len(str(celda.value)) > longitud_maxima:
                    longitud_maxima = len(celda.value)
            except:
                pass
        # Agrega un margen adicional y un factor de ajuste
        adjusted_width = (longitud_maxima + 2) * 1.2  
        hoja.column_dimensions[column].width = adjusted_width

    #Se crea la gráfica de barras con lo siguiente--------------------------------------------------------------------
    valores = Reference(hoja, min_col=4, min_row=5, max_row=len(productos) + 4)
    categorias = Reference(hoja, min_col=3, min_row=5, max_row=len(productos) + 4)

    # Crear el gráfico de barras
    grafica = BarChart()
    grafica.add_data(valores)  # Agregar datos con títulos de categorías
    grafica.set_categories(categorias)

    # Agregar el gráfico a la hoja de cálculo
    hoja.add_chart(grafica, f"D{ posicionGrafica }")

    # Guardar el libro
    libro.save('prueba.xlsx')
    libro.close()

crearDocumento()
